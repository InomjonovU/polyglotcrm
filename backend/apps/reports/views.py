import io
from datetime import date, timedelta
from decimal import Decimal
from django.db.models import Sum, Count, Avg, Q
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.users.permissions import IsAdmin
from apps.students.models import Student, StudentStatus
from apps.groups.models import Group
from apps.teachers.models import Teacher
from apps.payments.models import Payment, MonthlyCharge
from apps.salary.models import SalaryRecord
from apps.attendance.models import Attendance
from apps.grades.models import Grade


@api_view(['GET'])
@permission_classes([IsAdmin])
def dashboard(request):
    # Oy almashganda avtomatik oylik hisoblarni yaratish
    try:
        from apps.payments.services import ensure_current_month_charges
        ensure_current_month_charges()
    except Exception:
        pass
    today = date.today()
    total_students = Student.objects.filter(status=StudentStatus.ACTIVE).count()
    charges = MonthlyCharge.objects.filter(year=today.year, month=today.month)
    total_charged = charges.aggregate(s=Sum('amount'))['s'] or 0
    total_paid = Payment.objects.filter(
        charge__year=today.year, charge__month=today.month
    ).aggregate(s=Sum('amount'))['s'] or 0
    paid_pct = round((total_paid / total_charged * 100) if total_charged else 0, 1)
    # 30 kunlik daromad
    last30 = []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        s = Payment.objects.filter(paid_at__date=d).aggregate(s=Sum('amount'))['s'] or 0
        last30.append({'date': d.isoformat(), 'amount': float(s)})

    return Response({
        'total_students': total_students,
        'groups_count': Group.objects.filter(is_active=True).count(),
        'teachers_count': Teacher.objects.filter(is_active=True).count(),
        'total_charged': float(total_charged),
        'total_paid': float(total_paid),
        'paid_percent': paid_pct,
        'balance': float(Decimal(total_charged) - Decimal(total_paid)),
        'income_last_30_days': last30,
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def financial(request):
    year = int(request.query_params.get('year', date.today().year))
    monthly = []
    for m in range(1, 13):
        charged = MonthlyCharge.objects.filter(year=year, month=m).aggregate(s=Sum('amount'))['s'] or 0
        paid = Payment.objects.filter(charge__year=year, charge__month=m).aggregate(s=Sum('amount'))['s'] or 0
        salary = SalaryRecord.objects.filter(year=year, month=m).aggregate(s=Sum('paid_amount'))['s'] or 0
        monthly.append({
            'month': m,
            'charged': float(charged),
            'paid': float(paid),
            'salary': float(salary),
            'net': float(Decimal(paid) - Decimal(salary)),
        })
    return Response({'year': year, 'monthly': monthly})


@api_view(['GET'])
@permission_classes([IsAdmin])
def teachers_report(request):
    result = []
    for t in Teacher.objects.select_related('user').filter(is_active=True):
        lessons = Attendance.objects.filter(group__teacher=t).count()
        students_stats = Attendance.objects.filter(group__teacher=t).aggregate(
            present=Count('id', filter=Q(status='present')),
            total=Count('id'),
        )
        avg_percent = 0
        if students_stats['total']:
            avg_percent = round(students_stats['present'] * 100 / students_stats['total'], 1)
        avg_grade = Grade.objects.filter(group__teacher=t).aggregate(a=Avg('value'))['a']
        result.append({
            'id': t.id,
            'name': t.user.full_name,
            'groups_count': t.groups.filter(is_active=True).count(),
            'attendance_entries': lessons,
            'avg_attendance_percent': avg_percent,
            'avg_grade': round(avg_grade, 2) if avg_grade else None,
        })
    return Response(result)


# ============================================================
# Excel eksportlar
# ============================================================

def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, ncols):
    """Bosh qatorni stillaymiz — qora fon, oq matn, bold."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill('solid', fgColor='1F2937')
    font = Font(bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    side = Side(style='thin', color='CBD5E1')
    border = Border(left=side, right=side, top=side, bottom=side)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'


def _autosize(ws, max_width=42):
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        m = 8
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            ln = len(str(v))
            if ln > m:
                m = ln
        ws.column_dimensions[get_column_letter(col_idx)].width = min(m + 2, max_width)


@api_view(['GET'])
@permission_classes([IsAdmin])
def export_students(request):
    """Barcha o'quvchilar Excel fayl sifatida — barcha ma'lumotlari bilan."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    headers = [
        '#', 'Ism', 'Familiya', 'Telefon', 'Tug\'ilgan sana', 'Manzil',
        'Holat', 'Chegirma %', 'Qo\'shilgan sana',
        'Asosiy guruh', 'Barcha guruhlar',
        'Ota-ona ismi', 'Ota-ona telefoni',
        "Kim taklif qildi", 'Tashqi manba', 'Izoh',
    ]
    ws.append(headers)

    qs = Student.objects.select_related('user', 'group').prefetch_related(
        'groups', 'referrer_student__user', 'referrer_teacher__user'
    ).order_by('user__first_name', 'user__last_name')

    status_map = {
        StudentStatus.ACTIVE: 'Faol',
        StudentStatus.FROZEN: 'Muzlatilgan',
        StudentStatus.ARCHIVED: 'Arxiv',
    }

    for i, s in enumerate(qs, start=1):
        u = s.user
        groups = ', '.join(g.name for g in s.groups.all())
        if s.referrer_student_id and s.referrer_student.user:
            referrer = f"O'quvchi: {s.referrer_student.user.full_name}"
        elif s.referrer_teacher_id and s.referrer_teacher.user:
            referrer = f"O'qituvchi: {s.referrer_teacher.user.full_name}"
        else:
            referrer = ''
        ws.append([
            i,
            u.first_name,
            u.last_name,
            u.phone or '',
            u.birth_date.isoformat() if u.birth_date else '',
            u.address or '',
            status_map.get(s.status, s.status),
            float(s.discount_percent or 0),
            s.joined_date.isoformat() if s.joined_date else '',
            s.group.name if s.group else '',
            groups,
            s.parent_name or '',
            s.parent_phone or '',
            referrer,
            s.referrer_source or '',
            s.notes or '',
        ])

    _style_header(ws, len(headers))
    _autosize(ws)

    fname = f"oquvchilar_{date.today().isoformat()}.xlsx"
    return _xlsx_response(wb, fname)


@api_view(['GET'])
@permission_classes([IsAdmin])
def export_payments(request):
    """Barcha to'lovlar tarixi Excel fayl sifatida."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "To'lovlar"

    headers = [
        '#', 'Sana', 'Vaqt', 'Chek kodi', 'O\'quvchi', 'Telefon',
        'Guruh', 'Oy', 'Yil', 'Summa', 'To\'lov turi', 'Qabul qildi', 'Izoh',
    ]
    ws.append(headers)

    method_map = {
        'cash': 'Naqd',
        'card': 'Plastik',
        'transfer': "Bank o'tkazma",
    }

    qs = Payment.objects.select_related(
        'student__user', 'charge__group', 'received_by'
    ).order_by('-paid_at')

    # Filter ixtiyoriy: ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
    df = request.query_params.get('date_from')
    dt = request.query_params.get('date_to')
    if df:
        qs = qs.filter(paid_at__date__gte=df)
    if dt:
        qs = qs.filter(paid_at__date__lte=dt)

    for i, p in enumerate(qs, start=1):
        paid = p.paid_at
        u = p.student.user if p.student_id else None
        group_name = p.charge.group.name if (p.charge_id and p.charge and p.charge.group) else ''
        rb = p.received_by.full_name if p.received_by else ''
        ws.append([
            i,
            paid.date().isoformat() if paid else '',
            paid.strftime('%H:%M') if paid else '',
            p.receipt_code or '',
            u.full_name if u else '',
            u.phone if u else '',
            group_name,
            p.charge.month if p.charge_id and p.charge else '',
            p.charge.year if p.charge_id and p.charge else '',
            float(p.amount or 0),
            method_map.get(p.method, p.method),
            rb,
            p.note or '',
        ])

    # Pastda jami summa
    total = qs.aggregate(s=Sum('amount'))['s'] or 0
    ws.append([])
    last = ws.max_row + 1
    ws.cell(row=last, column=9, value='JAMI:')
    cell = ws.cell(row=last, column=10, value=float(total))
    from openpyxl.styles import Font
    ws.cell(row=last, column=9).font = Font(bold=True)
    cell.font = Font(bold=True)

    _style_header(ws, len(headers))
    _autosize(ws)

    fname = f"tolovlar_{date.today().isoformat()}.xlsx"
    return _xlsx_response(wb, fname)


@api_view(['GET'])
@permission_classes([IsAdmin])
def students_dynamics(request):
    year = int(request.query_params.get('year', date.today().year))
    monthly = []
    for m in range(1, 13):
        joined = Student.objects.filter(joined_date__year=year, joined_date__month=m).count()
        archived = Student.objects.filter(status=StudentStatus.ARCHIVED).count()  # approximate
        frozen = Student.objects.filter(status=StudentStatus.FROZEN).count()
        monthly.append({'month': m, 'new': joined, 'archived': archived, 'frozen': frozen})
    groups = []
    for g in Group.objects.filter(is_active=True):
        cnt = g.enrolled_students.filter(status=StudentStatus.ACTIVE).count()
        cap = g.max_students or cnt
        groups.append({
            'id': g.id,
            'name': g.name,
            'active': cnt,
            'capacity': cap,
            'fill_percent': round(cnt * 100 / cap, 1) if cap else 100,
        })
    return Response({'monthly': monthly, 'groups': groups})
